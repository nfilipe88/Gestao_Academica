"""
Exportação do relatório de Estatísticas (ver app/cruds/estatisticas.py)
para Excel, em dois formatos — pedido explicitamente:

- .xlsx (openpyxl): formato moderno, o que a maioria dos utilizadores
  realmente quer.
- .xls (xlwt): formato binário pré-2007, mantido só para quem tem
  mesmo um sistema/fluxo que só aceita este formato antigo. xlwt está
  sem manutenção há anos mas continua a funcionar bem para exports
  simples e sem formatação complexa como este.

Ambos os formatos partilham a mesma estrutura de folhas
(_construir_folhas), montada uma única vez a partir do dict devolvido
por cruds/estatisticas.py::obter_relatorio — só a serialização final
(openpyxl vs xlwt) é diferente.
"""
import io
from decimal import Decimal

import openpyxl
import xlwt
from openpyxl.styles import Font


def _construir_folhas(relatorio: dict) -> list[tuple[str, list[str], list[list]]]:
    """Devolve [(nome_da_folha, cabecalhos, linhas), ...] — os valores
    já vêm prontos a escrever (Decimal/date funcionam nos dois
    escritores, não precisam de conversão extra aqui)."""
    resumo_financeiro_atraso = relatorio["atrasos_periodo"]

    return [
        ("Resumo", ["Indicador", "Valor"], [
            ["Período", f"{relatorio['data_inicio']} a {relatorio['data_fim']}"],
            ["Total de alunos matriculados", relatorio["total_alunos_matriculados"]],
            ["Matrículas no período", relatorio["matriculas_no_periodo"]],
            ["Total de entradas no período", relatorio["total_entradas_periodo"]],
            ["Total de saídas no período", relatorio["total_saidas_periodo"]],
            ["Saldo do período", relatorio["saldo_periodo"]],
            ["Faturas atrasadas no período", resumo_financeiro_atraso["total_faturas_atrasadas"]],
            ["Valor total em atraso", resumo_financeiro_atraso["valor_total_atraso"]],
        ]),
        ("Faixas Etárias", ["Faixa Etária", "Total de Alunos"], [
            [f["faixa"], f["total"]] for f in relatorio["faixas_etarias"]
        ]),
        ("Cursos Mais Concorridos", ["Curso", "Alunos Matriculados"], [
            [c["nome_curso"], c["total_matriculados"]] for c in relatorio["cursos_mais_concorridos"]
        ]),
        ("Disciplinas — Aproveitamento", ["Disciplina", "Média", "Nº de Notas"], [
            [d["nome_disciplina"], d["media"], d["total_notas"]] for d in relatorio["disciplinas_maior_aproveitamento"]
        ]),
        ("Turmas — Melhores Notas", ["Turma", "Média", "Alunos Avaliados"], [
            [t["nome_turma"], t["media"], t["total_alunos_avaliados"]] for t in relatorio["turmas_melhores_notas"]
        ]),
        ("Alunos — Melhores Notas", ["Aluno", "Média", "Nº de Notas"], [
            [a["nome_aluno"], a["media"], a["total_notas"]] for a in relatorio["alunos_melhores_notas"]
        ]),
        ("Pagamentos por Mês", ["Mês", "Nº de Pagamentos", "Valor Total"], [
            [p["mes"], p["total_pagamentos"], p["valor_total"]] for p in relatorio["pagamentos_por_mes"]
        ]),
        ("Maiores Entradas", ["Aluno", "Valor", "Data do Pagamento", "Forma de Pagamento"], [
            [e["nome_aluno"], e["valor"], e["data_pagamento"], e["forma_pagamento"]] for e in relatorio["maiores_entradas"]
        ]),
        ("Despesas por Mês", ["Mês", "Nº de Despesas", "Valor Total"], [
            [d["mes"], d["total_despesas"], d["valor_total"]] for d in relatorio["despesas_por_mes"]
        ]),
        ("Maiores Saídas", ["Categoria", "Descrição", "Valor", "Data"], [
            [s["categoria"], s["descricao"], s["valor"], s["data_despesa"]] for s in relatorio["maiores_saidas"]
        ]),
    ]


def gerar_xlsx(relatorio: dict) -> bytes:
    livro = openpyxl.Workbook()
    livro.remove(livro.active)  # a folha em branco criada por omissão

    for nome_folha, cabecalhos, linhas in _construir_folhas(relatorio):
        folha = livro.create_sheet(title=nome_folha[:31])  # limite do Excel: 31 carateres no nome da folha
        folha.append(cabecalhos)
        for celula in folha[1]:
            celula.font = Font(bold=True)
        for linha in linhas:
            folha.append([_valor_serializavel(v) for v in linha])
        for coluna in folha.columns:
            largura = max((len(str(c.value)) for c in coluna if c.value is not None), default=10)
            folha.column_dimensions[coluna[0].column_letter].width = min(largura + 2, 50)

    buffer = io.BytesIO()
    livro.save(buffer)
    return buffer.getvalue()


def gerar_xls(relatorio: dict) -> bytes:
    livro = xlwt.Workbook(encoding="utf-8")
    estilo_cabecalho = xlwt.easyxf("font: bold on")

    for nome_folha, cabecalhos, linhas in _construir_folhas(relatorio):
        # xlwt: nome da folha limitado a 31 carateres, tal como .xlsx.
        folha = livro.add_sheet(nome_folha[:31])
        for col, cabecalho in enumerate(cabecalhos):
            folha.write(0, col, cabecalho, estilo_cabecalho)
        for linha_idx, linha in enumerate(linhas, start=1):
            for col, valor in enumerate(linha):
                folha.write(linha_idx, col, _valor_serializavel(valor))

    buffer = io.BytesIO()
    livro.save(buffer)
    return buffer.getvalue()


def _valor_serializavel(valor):
    """Decimal não é aceite diretamente por nenhum dos dois escritores
    — convertido para float só na escrita, nunca nos cálculos em
    cruds/estatisticas.py (que continuam em Decimal, para não perder
    precisão nas somas)."""
    if isinstance(valor, Decimal):
        return float(valor)
    if hasattr(valor, "isoformat"):  # date/datetime
        return valor.isoformat()
    return valor
