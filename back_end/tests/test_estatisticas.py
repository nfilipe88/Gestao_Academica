"""Estatísticas — Dashboard (estado corrente) e Relatório por
intervalo de datas, exportável em .xlsx/.xls (ver
app/cruds/estatisticas.py e app/core/estatisticas_excel.py). Módulo
novo, sem nenhum teste antes desta sessão."""
import io
from datetime import date, timedelta

import openpyxl

from tests.conftest import auth_headers, criar_escola_e_gestor, sufixo_unico
from tests.test_comportamento import _criar_professor_com_token


def _data_nascimento_para_idade(idade: int) -> str:
    """1 de janeiro é sempre "já passado" em relação a qualquer data de
    execução real dos testes — a idade calculada por
    cruds/estatisticas.py::_calcular_idade fica exatamente `idade`,
    sem depender do dia em que a suite corre."""
    return f"{date.today().year - idade}-01-01"


async def _criar_curso_serie_turma(client, headers, ano_letivo: int, nome_curso: str, nome_turma: str) -> str:
    resp = await client.post("/api/v1/academico/cursos", headers=headers, json={"nome": nome_curso})
    curso_id = resp.json()["id"]
    resp = await client.post("/api/v1/academico/series", headers=headers, json={"curso_id": curso_id, "nome": "Série Única"})
    serie_id = resp.json()["id"]
    resp = await client.post("/api/v1/academico/turmas", headers=headers, json={
        "serie_ano_id": serie_id, "nome_codigo": nome_turma, "ano_letivo": ano_letivo, "vagas_maximas": 30
    })
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


async def _criar_aluno_matriculado(client, headers, turma_id: str, ano_letivo: int, nome: str, idade: int) -> dict:
    suf = sufixo_unico()
    resp = await client.post("/api/v1/alunos", headers=headers, json={
        "matricula_interna": f"AL{suf}", "nome_completo": nome, "data_nascimento": _data_nascimento_para_idade(idade)
    })
    aluno_id = resp.json()["id"]
    resp = await client.post("/api/v1/matriculas", headers=headers, json={
        "aluno_id": aluno_id, "turma_id": turma_id, "ano_letivo": ano_letivo
    })
    assert resp.status_code == 201, resp.text
    return {"aluno_id": aluno_id, "matricula_id": resp.json()["id"], "nome": nome}


async def _lancar_nota(client, headers, turma_id, disciplina_id, matricula_id, valor: str, data_avaliacao: str):
    resp = await client.post(
        f"/api/v1/diario/turmas/{turma_id}/disciplinas/{disciplina_id}/notas/lote", headers=headers, json={
            "periodo_avaliacao": "1º Trimestre", "data_avaliacao": data_avaliacao,
            "notas": [{"matricula_id": matricula_id, "valor_nota": valor}]
        }
    )
    assert resp.status_code == 200, resp.text


async def _montar_cenario(client, headers) -> dict:
    """5 alunos em 2 turmas de 2 cursos diferentes, cobrindo as 4
    faixas etárias (uma delas com 2 alunos), com notas lançadas para
    dar um ranking inequívoco de disciplina/turma/aluno."""
    ano_letivo = date.today().year
    hoje = date.today()

    turma_a_id = await _criar_curso_serie_turma(client, headers, ano_letivo, "Ensino Secundário", "10º A")
    turma_b_id = await _criar_curso_serie_turma(client, headers, ano_letivo, "Ensino Básico", "5º A")

    resp = await client.post("/api/v1/academico/disciplinas", headers=headers, json={
        "nome": f"Matemática {sufixo_unico()}", "carga_horaria_total": 4
    })
    disciplina_id = resp.json()["id"]

    aluno_7 = await _criar_aluno_matriculado(client, headers, turma_a_id, ano_letivo, "Aluno Sete Anos", 7)
    aluno_12 = await _criar_aluno_matriculado(client, headers, turma_a_id, ano_letivo, "Aluno Doze Anos", 12)
    aluno_16 = await _criar_aluno_matriculado(client, headers, turma_a_id, ano_letivo, "Aluno Dezasseis Anos", 16)
    aluno_20 = await _criar_aluno_matriculado(client, headers, turma_b_id, ano_letivo, "Aluno Vinte Anos", 20)
    aluno_13 = await _criar_aluno_matriculado(client, headers, turma_b_id, ano_letivo, "Aluno Treze Anos", 13)

    # Ranking: aluno_16 > aluno_12 > aluno_7 (turma A); aluno_13 > aluno_20 (turma B).
    # Turma A fica com a média mais alta (9.5/8.0/4.0 = 7.17) vs Turma B (7.5/5.0 = 6.25).
    await _lancar_nota(client, headers, turma_a_id, disciplina_id, aluno_16["matricula_id"], "9.5", str(hoje))
    await _lancar_nota(client, headers, turma_a_id, disciplina_id, aluno_12["matricula_id"], "8.0", str(hoje))
    await _lancar_nota(client, headers, turma_a_id, disciplina_id, aluno_7["matricula_id"], "4.0", str(hoje))
    await _lancar_nota(client, headers, turma_b_id, disciplina_id, aluno_13["matricula_id"], "7.5", str(hoje))
    await _lancar_nota(client, headers, turma_b_id, disciplina_id, aluno_20["matricula_id"], "5.0", str(hoje))

    # Financeiro: um contrato para o aluno_16, parcela 1 marcada paga hoje.
    resp = await client.post("/api/v1/responsaveis", headers=headers, json={
        "nome_completo": "Responsável Estatísticas", "telefone_contato": "+244900000000"
    })
    responsavel_id = resp.json()["id"]
    resp = await client.post(f"/api/v1/alunos/{aluno_16['aluno_id']}/responsaveis", headers=headers, json={
        "responsavel_id": responsavel_id, "tipo_parentesco": "Mãe", "responsavel_financeiro": True
    })
    assert resp.status_code == 201, resp.text
    resp = await client.post("/api/v1/financeiro/contratos", headers=headers, json={
        "matricula_id": aluno_16["matricula_id"], "responsavel_id": responsavel_id,
        "valor_total_anual": "1200.00", "quantidade_parcelas": 12
    })
    assert resp.status_code == 201, resp.text
    contrato_id = resp.json()["id"]
    resp = await client.get(f"/api/v1/financeiro/contratos/{contrato_id}/faturas", headers=headers)
    primeira_fatura_id = next(f["id"] for f in resp.json() if f["numero_parcela"] == 1)
    resp = await client.patch(f"/api/v1/financeiro/faturas/{primeira_fatura_id}/marcar-pago", headers=headers, json={
        "valor_pago": "100.00", "forma_pagamento": "MANUAL"
    })
    assert resp.status_code == 200, resp.text

    # Despesa lançada hoje.
    await client.post("/api/v1/financeiro/despesas", headers=headers, json={
        "categoria": "RENDA", "descricao": "Renda do mês", "valor": "300.00", "data_despesa": str(hoje)
    })

    return {
        "turma_a_id": turma_a_id, "turma_b_id": turma_b_id, "disciplina_id": disciplina_id,
        "aluno_7": aluno_7, "aluno_12": aluno_12, "aluno_16": aluno_16, "aluno_20": aluno_20, "aluno_13": aluno_13,
        "hoje": hoje,
    }


async def test_dashboard_total_matriculados_e_faixas_etarias(client):
    escola = await criar_escola_e_gestor(client, "estatisticas-dashboard-basico")
    headers = auth_headers(escola["token"])
    await _montar_cenario(client, headers)

    resp = await client.get("/api/v1/estatisticas/dashboard", headers=headers)
    assert resp.status_code == 200, resp.text
    dados = resp.json()

    assert dados["total_alunos_matriculados"] == 5
    faixas = {f["faixa"]: f["total"] for f in dados["faixas_etarias"]}
    assert faixas == {"5-9": 1, "10-14": 2, "15-18": 1, "19+": 1}


async def test_dashboard_cursos_mais_concorridos(client):
    escola = await criar_escola_e_gestor(client, "estatisticas-cursos-concorridos")
    headers = auth_headers(escola["token"])
    await _montar_cenario(client, headers)

    resp = await client.get("/api/v1/estatisticas/dashboard", headers=headers)
    cursos = resp.json()["cursos_mais_concorridos"]
    assert cursos[0]["nome_curso"] == "Ensino Secundário"
    assert cursos[0]["total_matriculados"] == 3
    assert cursos[1]["nome_curso"] == "Ensino Básico"
    assert cursos[1]["total_matriculados"] == 2


async def test_dashboard_ranking_disciplinas_turmas_alunos(client):
    escola = await criar_escola_e_gestor(client, "estatisticas-ranking-notas")
    headers = auth_headers(escola["token"])
    cenario = await _montar_cenario(client, headers)

    resp = await client.get("/api/v1/estatisticas/dashboard", headers=headers)
    dados = resp.json()

    assert dados["disciplinas_maior_aproveitamento"][0]["total_notas"] == 5
    assert abs(dados["disciplinas_maior_aproveitamento"][0]["media"] - 6.8) < 0.05

    assert dados["turmas_melhores_notas"][0]["turma_id"] == cenario["turma_a_id"]
    assert dados["turmas_melhores_notas"][1]["turma_id"] == cenario["turma_b_id"]

    assert dados["alunos_melhores_notas"][0]["nome_aluno"] == "Aluno Dezasseis Anos"
    assert dados["alunos_melhores_notas"][0]["media"] == 9.5


async def test_relatorio_matriculas_no_periodo(client):
    escola = await criar_escola_e_gestor(client, "estatisticas-matriculas-periodo")
    headers = auth_headers(escola["token"])
    await _montar_cenario(client, headers)
    hoje = date.today()

    resp = await client.get(
        f"/api/v1/estatisticas/relatorio?data_inicio={hoje}&data_fim={hoje}", headers=headers
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["matriculas_no_periodo"] == 5

    resp = await client.get(
        f"/api/v1/estatisticas/relatorio?data_inicio={hoje - timedelta(days=365)}&data_fim={hoje - timedelta(days=30)}",
        headers=headers
    )
    assert resp.json()["matriculas_no_periodo"] == 0


async def test_relatorio_financeiro_entradas_saidas_e_saldo(client):
    escola = await criar_escola_e_gestor(client, "estatisticas-financeiro-periodo")
    headers = auth_headers(escola["token"])
    await _montar_cenario(client, headers)
    hoje = date.today()

    resp = await client.get(
        f"/api/v1/estatisticas/relatorio?data_inicio={hoje}&data_fim={hoje}", headers=headers
    )
    dados = resp.json()

    assert len(dados["pagamentos_por_mes"]) == 1
    assert dados["pagamentos_por_mes"][0]["total_pagamentos"] == 1
    assert float(dados["pagamentos_por_mes"][0]["valor_total"]) == 100.0

    assert len(dados["maiores_entradas"]) == 1
    assert dados["maiores_entradas"][0]["nome_aluno"] == "Aluno Dezasseis Anos"

    assert len(dados["despesas_por_mes"]) == 1
    assert float(dados["despesas_por_mes"][0]["valor_total"]) == 300.0

    assert len(dados["maiores_saidas"]) == 1
    assert dados["maiores_saidas"][0]["categoria"] == "RENDA"

    assert float(dados["total_entradas_periodo"]) == 100.0
    assert float(dados["total_saidas_periodo"]) == 300.0
    assert float(dados["saldo_periodo"]) == -200.0


async def test_relatorio_atrasos_periodo_estrutura_e_zero_quando_nada_vencido(client):
    """A deteção de "atrasado" em si (RN02, calcular_situacao_fatura) já
    é testada em test_matricula_financeiro.py/test_recibo_pagamento.py;
    aqui cobre-se só a agregação: um contrato recém-criado não tem
    nada vencido, logo 0 atrasos, com a estrutura correta."""
    escola = await criar_escola_e_gestor(client, "estatisticas-atrasos")
    headers = auth_headers(escola["token"])
    await _montar_cenario(client, headers)

    hoje = date.today()
    resp = await client.get(
        f"/api/v1/estatisticas/relatorio?data_inicio={hoje - timedelta(days=60)}&data_fim={hoje + timedelta(days=60)}",
        headers=headers
    )
    assert resp.status_code == 200, resp.text
    atrasos = resp.json()["atrasos_periodo"]
    assert atrasos["total_faturas_atrasadas"] == 0
    assert float(atrasos["valor_total_atraso"]) == 0.0


async def test_relatorio_data_fim_antes_de_inicio_e_rejeitado(client):
    escola = await criar_escola_e_gestor(client, "estatisticas-datas-invalidas")
    headers = auth_headers(escola["token"])
    hoje = date.today()

    resp = await client.get(
        f"/api/v1/estatisticas/relatorio?data_inicio={hoje}&data_fim={hoje - timedelta(days=1)}", headers=headers
    )
    assert resp.status_code == 400, resp.text


async def test_estatisticas_professor_bloqueado(client):
    escola = await criar_escola_e_gestor(client, "estatisticas-rbac-professor")
    headers = auth_headers(escola["token"])
    _, token_professor = await _criar_professor_com_token(client, headers, "Prof. Estatísticas")
    headers_professor = auth_headers(token_professor)

    resp = await client.get("/api/v1/estatisticas/dashboard", headers=headers_professor)
    assert resp.status_code == 403, resp.text


async def test_estatisticas_isoladas_por_tenant(client):
    escola_a = await criar_escola_e_gestor(client, "estatisticas-iso-a")
    escola_b = await criar_escola_e_gestor(client, "estatisticas-iso-b")
    headers_a = auth_headers(escola_a["token"])
    headers_b = auth_headers(escola_b["token"])
    await _montar_cenario(client, headers_a)

    resp = await client.get("/api/v1/estatisticas/dashboard", headers=headers_b)
    assert resp.status_code == 200, resp.text
    assert resp.json()["total_alunos_matriculados"] == 0


async def test_exportar_xlsx_contem_as_folhas_e_dados_esperados(client):
    escola = await criar_escola_e_gestor(client, "estatisticas-export-xlsx")
    headers = auth_headers(escola["token"])
    await _montar_cenario(client, headers)
    hoje = date.today()

    resp = await client.get(
        f"/api/v1/estatisticas/relatorio.xlsx?data_inicio={hoje}&data_fim={hoje}", headers=headers
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    livro = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Resumo" in livro.sheetnames
    assert "Faixas Etárias" in livro.sheetnames
    assert "Maiores Saídas" in livro.sheetnames

    folha_faixas = livro["Faixas Etárias"]
    linhas = {row[0]: row[1] for row in folha_faixas.iter_rows(min_row=2, values_only=True)}
    assert linhas["5-9"] == 1
    assert linhas["10-14"] == 2


async def test_exportar_xls_devolve_ficheiro_binario_valido(client):
    escola = await criar_escola_e_gestor(client, "estatisticas-export-xls")
    headers = auth_headers(escola["token"])
    await _montar_cenario(client, headers)
    hoje = date.today()

    resp = await client.get(
        f"/api/v1/estatisticas/relatorio.xls?data_inicio={hoje}&data_fim={hoje}", headers=headers
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/vnd.ms-excel"
    # Assinatura binária de um ficheiro OLE2/.xls (o mesmo container do .doc/.ppt antigos).
    assert resp.content.startswith(b"\xd0\xcf\x11\xe0")
